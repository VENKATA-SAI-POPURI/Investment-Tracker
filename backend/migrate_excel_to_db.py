"""
Migrate data from 'Investment Listing.xlsx' into investments.db (SQLite).
Run once:  python migrate_excel_to_db.py
"""
import os
import sys
from openpyxl import load_workbook
from db_service import DbService, SHEET_TO_TABLE, TABLES, NUMERIC_FIELDS

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "Investment Listing.xlsx")
DB_PATH = os.path.join(BASE_DIR, "backend", "investments.db")


def migrate():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    # Remove existing DB so we start fresh
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"Removed existing {DB_PATH}")

    db = DbService(DB_PATH)
    wb = load_workbook(EXCEL_PATH)

    total = 0
    for sheet_name, table in SHEET_TO_TABLE.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping {sheet_name} (not in workbook)")
            continue

        config = TABLES[table]
        columns = config["columns"]
        ws = wb[sheet_name]
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            data = {}
            for i, col in enumerate(columns):
                val = row[i] if i < len(row) else None
                if val is None:
                    data[col] = ""
                    continue
                if col in NUMERIC_FIELDS:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = 0
                elif hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val)
                data[col] = val
            # Use direct insert (bypass upsert logic)
            conn = db._connect()
            cols_str = ", ".join(columns)
            placeholders = ", ".join("?" for _ in columns)
            vals = [data.get(c, "") for c in columns]
            conn.execute(f"INSERT INTO {table} ({cols_str}) VALUES ({placeholders})", vals)
            conn.commit()
            conn.close()
            count += 1

        print(f"  {sheet_name}: {count} rows migrated")
        total += count

    wb.close()
    print(f"\nDone! {total} total rows migrated to {DB_PATH}")


if __name__ == "__main__":
    migrate()
