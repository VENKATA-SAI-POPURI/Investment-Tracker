import threading
import shutil
import os
from datetime import datetime
from openpyxl import load_workbook

NUMERIC_FIELDS = {
    "buy_quantity", "buy_value", "sell_quantity", "sell_value",
    "amount", "tenure", "fd_value",
    "interest", "return_value",
    "inr_amount", "usd_amount", "rate",
}

DATE_FIELDS = {"date", "maturity_date"}

SHEET_CONFIG = {
    "Equity": {
        "columns": ["year", "market", "market_cap", "sector", "name", "date", "buy_quantity", "buy_value", "sell_quantity", "sell_value", "buy_sell", "remarks"],
        "buy_col": "buy_value",
        "sell_col": "sell_value",
    },
    "Commodity": {
        "columns": ["year", "commodity", "name", "date", "buy_quantity", "buy_value", "sell_quantity", "sell_value", "buy_sell", "remarks"],
        "buy_col": "buy_value",
        "sell_col": "sell_value",
    },
    "Mutual Funds": {
        "columns": ["year", "category", "fund_type", "name", "date", "buy_quantity", "buy_value", "sell_quantity", "sell_value", "buy_sell", "remarks"],
        "buy_col": "buy_value",
        "sell_col": "sell_value",
    },
    "P2P": {
        "columns": ["lending_id", "platform", "name", "date", "amount", "tenure", "maturity_date", "status", "remarks"],
        "buy_col": "amount",
        "sell_col": None,
    },
    "P2P Repayments": {
        "columns": ["lending_id", "date", "amount", "remarks"],
        "buy_col": None,
        "sell_col": "amount",
    },
    "Fixed Deposits": {
        "columns": ["year", "platform", "bank_name", "date", "fd_value", "interest", "maturity_date", "return_value", "remarks"],
        "buy_col": "fd_value",
        "sell_col": "return_value",
    },
    "Forex": {
        "columns": ["date", "type", "inr_amount", "usd_amount", "rate", "remarks"],
        "buy_col": "inr_amount",
        "sell_col": "usd_amount",
    },
}


class ExcelService:
    def __init__(self, filepath):
        self.filepath = filepath
        self._lock = threading.Lock()
        self._backup_dir = os.path.join(os.path.dirname(filepath), "backups")
        os.makedirs(self._backup_dir, exist_ok=True)

    def _get_config(self, sheet_name):
        return SHEET_CONFIG[sheet_name]

    def _create_backup(self):
        """Create a timestamped backup of the Excel file before writing."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"Investment_Listing_backup_{timestamp}.xlsx"
        backup_path = os.path.join(self._backup_dir, backup_name)
        shutil.copy2(self.filepath, backup_path)
        # Keep only last 10 backups
        backups = sorted(
            [f for f in os.listdir(self._backup_dir) if f.endswith('.xlsx')],
            reverse=True
        )
        for old in backups[10:]:
            os.remove(os.path.join(self._backup_dir, old))

    def get_all(self, sheet_name):
        config = self._get_config(sheet_name)
        columns = config["columns"]
        with self._lock:
            wb = load_workbook(self.filepath)
            ws = wb[sheet_name]
            rows = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                values = [cell.value for cell in row]
                if all(v is None for v in values):
                    continue
                entry = {"id": idx}
                for col_key, cell in zip(columns, row):
                    val = cell.value
                    if col_key in DATE_FIELDS and val is not None:
                        val = val.strftime("%Y-%m-%d") if hasattr(val, "strftime") else str(val)
                    if col_key in NUMERIC_FIELDS and val is not None:
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            pass
                    entry[col_key] = val
                rows.append(entry)
            wb.close()
        return rows

    def _find_row_by_name(self, ws, columns, name):
        """Find existing row index by name (case-insensitive). Returns row number or None."""
        if "name" not in columns:
            # For Fixed Deposits, use bank_name as key
            name_key = "bank_name" if "bank_name" in columns else None
            if not name_key:
                return None
            name_col_idx = columns.index(name_key)
        else:
            name_col_idx = columns.index("name")

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if idx > ws.max_row:
                break
            cell_val = row[name_col_idx] if name_col_idx < len(row) else None
            if cell_val is not None and str(cell_val).strip().lower() == str(name).strip().lower():
                return idx
        return None

    def add_row(self, sheet_name, data):
        config = self._get_config(sheet_name)
        columns = config["columns"]

        # Determine the name key for this sheet
        name_key = "name" if "name" in columns else ("bank_name" if "bank_name" in columns else None)
        lookup_name = data.get(name_key, "").strip() if name_key else ""

        # Fields to update on upsert
        UPSERT_FIELDS = NUMERIC_FIELDS | DATE_FIELDS | {"buy_sell"}

        with self._lock:
            self._create_backup()
            wb = load_workbook(self.filepath)
            ws = wb[sheet_name]

            existing_row = self._find_row_by_name(ws, columns, lookup_name) if lookup_name else None

            if existing_row:
                # Update existing row: ADD numeric values, replace date/buy_sell
                for col_idx, col_key in enumerate(columns, start=1):
                    if col_key in UPSERT_FIELDS and col_key in data:
                        val = data[col_key]
                        if col_key in NUMERIC_FIELDS and val not in (None, ""):
                            try:
                                new_val = float(val)
                            except (ValueError, TypeError):
                                new_val = 0
                            existing_val = ws.cell(row=existing_row, column=col_idx).value
                            try:
                                existing_val = float(existing_val) if existing_val else 0
                            except (ValueError, TypeError):
                                existing_val = 0
                            val = existing_val + new_val
                        ws.cell(row=existing_row, column=col_idx, value=val)
                wb.save(self.filepath)
                wb.close()
                return {"id": existing_row, "upserted": True}
            else:
                row_values = []
                for col_key in columns:
                    val = data.get(col_key, "")
                    if col_key in NUMERIC_FIELDS and val not in (None, ""):
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            pass
                    row_values.append(val)
                ws.append(row_values)
                new_id = ws.max_row
                wb.save(self.filepath)
                wb.close()
                return {"id": new_id, "upserted": False}

    def update_row(self, sheet_name, row_id, data):
        config = self._get_config(sheet_name)
        columns = config["columns"]
        with self._lock:
            self._create_backup()
            wb = load_workbook(self.filepath)
            ws = wb[sheet_name]
            if row_id < 2 or row_id > ws.max_row:
                wb.close()
                return False
            for col_idx, col_key in enumerate(columns, start=1):
                if col_key in data:
                    val = data[col_key]
                    if col_key in NUMERIC_FIELDS and val not in (None, ""):
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            pass
                    ws.cell(row=row_id, column=col_idx, value=val)
            wb.save(self.filepath)
            wb.close()
        return True

    def delete_row(self, sheet_name, row_id):
        with self._lock:
            self._create_backup()
            wb = load_workbook(self.filepath)
            ws = wb[sheet_name]
            if row_id < 2 or row_id > ws.max_row:
                wb.close()
                return False
            ws.delete_rows(row_id)
            wb.save(self.filepath)
            wb.close()
        return True

    def get_summary(self):
        summary = {}
        with self._lock:
            wb = load_workbook(self.filepath)
            for sheet_name, config in SHEET_CONFIG.items():
                if sheet_name == "Forex":
                    continue
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                columns = config["columns"]
                buy_idx = columns.index(config["buy_col"]) if config["buy_col"] else None
                sell_idx = columns.index(config["sell_col"]) if config["sell_col"] else None
                total_buy = 0
                total_sell = 0
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    count += 1
                    buy_val = row[buy_idx] if buy_idx is not None and buy_idx < len(row) else None
                    sell_val = row[sell_idx] if sell_idx is not None and sell_idx < len(row) else None
                    if buy_val is not None:
                        try:
                            total_buy += float(buy_val)
                        except (ValueError, TypeError):
                            pass
                    if sell_val is not None:
                        try:
                            total_sell += float(sell_val)
                        except (ValueError, TypeError):
                            pass
                summary[sheet_name] = {
                    "count": count,
                    "total_buy": round(total_buy, 2),
                    "total_sell": round(total_sell, 2),
                    "net": round(total_sell - total_buy, 2),
                }
            wb.close()
        return summary
